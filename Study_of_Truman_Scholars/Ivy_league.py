import openpyxl

# Define variable to load dataframe
dataframe = openpyxl.load_workbook("winners.xlsx")

# Read the Excel file
dataframe1 = dataframe.active
Institutions = []
Ivy_League = ['Princeton University', 'Harvard University', 'Yale University', 'University of Pennsylvania',
              'Brown University', 'Columbia University ', 'Cornell University', 'Dartmouth College']
Princeton = 0
Harvard = 0
Yale = 0
UNP = 0
Brown = 0
Columbia = 0
Cornell = 0
Dartmouth = 0
count = 0

# # Find column with Institutions
# for i in range(1, dataframe1.max_column):
#     if dataframe1.cell(row=1, column=i).value == "Institution":
#         w = i

# Iterate the loop to read the cell values
for i in range(1, dataframe1.max_row):
    Institutions.append(dataframe1.cell(row=i, column=3).value)


for i in range(0, len(Institutions)):
    if Institutions[i] in Ivy_League:
        count += 1

    if Institutions[i] == 'Princeton University':
        Princeton += 1
    elif Institutions[i] == 'Harvard University':
        Harvard += 1
    elif Institutions[i] == 'Yale University':
        Yale += 1
    elif Institutions[i] == 'University of Pennsylvania':
         UNP += 1
    elif Institutions[i] == 'Brown University':
        Brown += 1
    elif Institutions[i] == 'Columbia University ':
        Columbia += 1
    elif Institutions[i] == 'Cornell University':
        Cornell += 1
    elif Institutions[i] == 'Dartmouth College':
        Dartmouth += 1


print("Number of Ivy League winners: " + str(count))
print("Princeton: " + str(Princeton))
print("Harvard: " + str(Harvard))
print("Yale: " + str(Yale))
print("UNP: " + str(UNP))
print("Brown: " + str(Brown))
print("Columbia: " + str(Columbia))
print("Cornell: " + str(Cornell))
print("Dartmouth: " + str(Dartmouth))
